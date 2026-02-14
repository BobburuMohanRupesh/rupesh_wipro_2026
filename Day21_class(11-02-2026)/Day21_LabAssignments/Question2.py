import pandas as pd
import openpyxl

# Step 1: Read Excel sheet into DataFrame
df = pd.read_excel("sales_data.xlsx", sheet_name="2025")
print(df)
# Step 2: Add new column Total
df["Total"] = df["Quantity"] * df["Price"]

# Step 3: Save updated DataFrame to new Excel file
df.to_excel("sales_summary.xlsx", index=False)

print("-----------------------------------------------------")
print("Sales summary saved successfully as sales_summary.xlsx")
df2 = pd.read_excel("sales_summary.xlsx")
print(df2)
#
# -------------------------------------------------------------
# USING PYXL
# ---------------------------------------------------------------
from openpyxl import load_workbook, Workbook

# Step 1: Load existing workbook
wb = load_workbook("sales_data.xlsx")

# Step 2: Select sheet "2025"
sheet = wb["2025"]

# Step 3: Create new workbook for output
new_wb = Workbook()
new_sheet = new_wb.active
new_sheet.title = "Sales Summary"

# Step 4: Copy headers + add Total column
headers = ["Product", "Quantity", "Price", "Total"]
new_sheet.append(headers)

# Step 5: Read rows from old sheet and calculate Total
for row in sheet.iter_rows(min_row=2, values_only=True):
    product, quantity, price = row
    total = quantity * price

    # Append updated row
    new_sheet.append([product, quantity, price, total])

# Step 6: Save new workbook
new_wb.save("sales_summary.xlsx")

print("Excel file created successfully using OpenPyXL")
